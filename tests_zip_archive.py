import os
from os.path import basename
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path, "test_archive.zip")


def test_archive_create():
    file_dir = os.listdir(path)
    with ZipFile(path_zip, "w") as zip_:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zip_.write(add_file, basename(add_file))
    assert zip_.namelist() == ['test_pdf.pdf', 'test_xlsx_10.xlsx', 'viikko.csv']
    os.remove('resources/test_archive.zip')


def test_csv():
    with ZipFile(path_zip) as archive:
        text = str(archive.read('viikko.csv'))
        assert text.__contains__('maanantai')


def test_xlsx():
    with ZipFile(path_zip) as archive:
        archive.extract('test_xlsx_10.xlsx')
        workbook = load_workbook('test_xlsx_10.xlsx')
        sheet = workbook.active
        check_value = str(sheet.cell(row=6, column=3).value)
        assert check_value == 'Magwood'
        os.remove('test_xlsx_10.xlsx')


def test_pdf():
    with ZipFile(path_zip) as archive:
        archive.extract('test_pdf.pdf')
        text = PdfReader('test_pdf.pdf').pages[0].extract_text()
        assert text.__contains__('Dumm')
        os.remove('test_pdf.pdf')